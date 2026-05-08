from openpyxl import load_workbook
wb = load_workbook("Scan.xlsx")
ws = wb.active

print("=== 检查 Scan.xlsx 输出 ===")
print(f"{'行':>3} | {'物品名称':20s} | {'数量':>6} | {'单位':>4} | {'含税单价':>10} | {'金额':>10} | {'备注':15s}")
print("-" * 90)

total = 0
bad_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    vals = [c.value for c in row]
    # Skip header/title rows
    if vals[0] and isinstance(vals[0], str) and ('物品名称' in vals[0] or '购' in vals[0] or '合同' in vals[0] or '合计' in vals[0]):
        print(f"  >> {vals}")
        continue
    if vals[0] is None and vals[1] is None:
        continue
    
    name = str(vals[0] or '') + str(vals[1] or '')
    qty = vals[2]
    unit = vals[3]
    price = vals[4]
    amount = vals[5]
    remark = vals[6] or ''
    
    # Check math
    ok = ""
    if qty and price and amount:
        try:
            expected = float(qty) * float(price)
            if abs(expected - float(amount)) > 0.1:
                ok = f" ❌ 应为{expected:.2f}"
                bad_rows.append((name[:15], qty, price, amount, expected))
        except:
            ok = " ⚠️ 无法验证"
    elif amount:
        try:
            total += float(amount)
        except:
            pass
        ok = " (缺数量/单价)"
    
    if qty and price and amount:
        try:
            total += float(amount)
        except:
            pass
    
    print(f"     | {name[:20]:20s} | {str(qty):>6} | {str(unit):>4} | {str(price):>10} | {str(amount):>10} | {remark[:15]:15s}{ok}")

print(f"\n=== 合计金额: {total:.2f} ===")
print(f"\n=== 有 {len(bad_rows)} 行数量*单价 ≠ 金额 ===")
for b in bad_rows[:20]:
    print(f"  {b[0]}: 数量={b[1]}, 单价={b[2]}, 金额={b[3]}, 应为={b[4]:.2f}")
