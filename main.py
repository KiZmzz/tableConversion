"""
PDF 表格转 Excel 工具（支持扫描件）
使用 PaddleOCR 进行中文 OCR + 坐标聚类提取表格

依赖: pip install pdfplumber openpyxl paddleocr paddlepaddle opencv-python-headless
用法: python main.py [PDF文件] [-o 输出文件]
"""

import os, sys, re, argparse
from pathlib import Path
import threading
import tkinter as tk
from tkinter import filedialog
import customtkinter as ctk

import cv2
import numpy as np
import pdfplumber
from paddleocr import PaddleOCR
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def detect_h_lines(img_path, min_len=60):
    """检测水平线以确定行位置"""
    img = cv2.imread(img_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_len, 1))
    h_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
    h_sum = np.sum(h_lines, axis=1)
    threshold = max(np.max(h_sum) * 0.3, 1)
    positions = np.where(h_sum > threshold)[0]

    rows = []
    group = [positions[0]] if len(positions) > 0 else []
    for p in positions[1:]:
        if p - group[-1] <= 8:
            group.append(p)
        else:
            rows.append(int(np.mean(group)))
            group = [p]
    if group:
        rows.append(int(np.mean(group)))
    return rows


def full_page_ocr(ocr_engine, img_path, engine_type='paddle'):
    """整页 OCR，返回带位置信息的结果"""
    items = []
    if engine_type == 'rapid':
        result, _ = ocr_engine(img_path)
        if not result: return items
        for line in result:
            box = line[0]
            text = line[1]
            conf = line[2]
            y_center = (float(box[0][1]) + float(box[2][1])) / 2
            x_center = (float(box[0][0]) + float(box[2][0])) / 2
            items.append({
                'text': text.strip(),
                'y': y_center,
                'x': x_center,
                'conf': float(conf)
            })
    else:
        result = ocr_engine.predict(img_path)
        if not result: return items
        for res in result:
            if 'rec_texts' not in res:
                continue
            for i in range(len(res['rec_texts'])):
                text = res['rec_texts'][i].strip()
                conf = float(res['rec_scores'][i])
                box = res['dt_polys'][i]  # [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
                y_center = (float(box[0][1]) + float(box[2][1])) / 2
                x_center = (float(box[0][0]) + float(box[2][0])) / 2
                items.append({
                    'text': text,
                    'y': y_center,
                    'x': x_center,
                    'conf': conf
                })
    return items


def assign_to_rows(items, h_lines):
    """将 OCR 文本块分配到最近的表格行"""
    if not h_lines or not items:
        return {}

    # 行区间：h_lines[i] ~ h_lines[i+1]
    rows = {}
    for item in items:
        y = item['y']
        row_idx = None
        for i in range(len(h_lines) - 1):
            if h_lines[i] <= y <= h_lines[i + 1]:
                row_idx = i
                break
        if row_idx is None:
            # 在最后一条线之下
            if y > h_lines[-1]:
                continue
            # 在第一条线之上
            if y < h_lines[0]:
                continue
        if row_idx is not None:
            if row_idx not in rows:
                rows[row_idx] = []
            rows[row_idx].append(item)
    return rows

def assign_to_columns(row_items, col_boundaries):
    """将一行中的文本块分配到列"""
    cells = {}
    for item in row_items:
        x = item['x']
        text = item['text'].strip()
        # 处理类似 "根根根根" 的 OCR 叠字错误
        if len(text) > 1 and len(set(text.replace(' ', ''))) == 1:
            text = text.replace(' ', '')[0]
            
        col_idx = len(col_boundaries) - 2  # 默认最后一列
        for c in range(len(col_boundaries) - 1):
            if col_boundaries[c] <= x < col_boundaries[c + 1]:
                col_idx = c
                break

        if col_idx in cells:
            # 同列合并：按 x 排序
            existing = cells[col_idx]
            # 去重：如果识别到的文本完全一致（例如合并单元格导致的重复），则不再重复添加
            if text == existing['text'].strip() or text in existing['text'].split():
                pass
            elif item['x'] < existing['x']:
                cells[col_idx] = {
                    'text': text + ' ' + existing['text'],
                    'x': item['x']
                }
            else:
                cells[col_idx] = {
                    'text': existing['text'] + ' ' + text,
                    'x': existing['x']
                }
        else:
            cells[col_idx] = {'text': text, 'x': item['x']}

    return {k: v['text'] for k, v in cells.items()}


def clean_number(s):
    """清洗 OCR 识别出的数字字符串，修正常见 OCR 错误
    返回 (float_value, cleaned_str) 或 (None, original_str)
    """
    if not s or not isinstance(s, str):
        return None, str(s) if s else ''
    s = s.strip()
    if not s:
        return None, ''
    
    # 常见 OCR 替换：冒号→点，中文逗号→空，空格→空
    cleaned = s.replace('：', '.').replace(':', '.').replace('，', '').replace(',', '').replace(' ', '')
    # 移除多余的点（例如 "1..16" → "1.16", "6. 5." → "6.5"）
    # 先处理连续两个点
    while '..' in cleaned:
        cleaned = cleaned.replace('..', '.')
    # 移除末尾的点
    cleaned = cleaned.rstrip('.')
    # 如果有多个小数点，只保留第一个
    parts = cleaned.split('.')
    if len(parts) > 2:
        cleaned = parts[0] + '.' + ''.join(parts[1:])
    
    # 移除非数字字符（除了小数点和负号）
    result = ''
    for ch in cleaned:
        if ch.isdigit() or ch == '.' or (ch == '-' and not result):
            result += ch
    
    if not result or result == '.' or result == '-':
        return None, s
    
    try:
        val = float(result)
        return val, result
    except ValueError:
        return None, s


def post_process_rows(rows, log_func=None):
    """对提取的表格行进行后处理：
    1. 清洗所有数值字段中的 OCR 错误（冒号→点等客观修正）
    2. 缺失值：仅当只缺一个值时，用另外两个推算补充
    3. 交叉校验 数量×单价=金额，不匹配的行标记为可疑（不自动修正）
    返回 suspicious_rows: set of row indices that need manual review
    """
    def log(msg):
        if log_func:
            log_func(msg)
    
    suspicious = set()
    fixed_count = 0
    
    for i, row in enumerate(rows):
        qty_raw = str(row.get(2, '')).strip()
        price_raw = str(row.get(4, '')).strip()
        amount_raw = str(row.get(5, '')).strip()
        
        qty_val, qty_clean = clean_number(qty_raw)
        price_val, price_clean = clean_number(price_raw)
        amount_val, amount_clean = clean_number(amount_raw)
        
        # 用清洗后的值更新（只是修正格式，如冒号→点）
        if qty_val is not None:
            row[2] = qty_clean
        if price_val is not None:
            row[4] = price_clean
        if amount_val is not None:
            row[5] = amount_clean
        
        # 统计有几个值是有效的
        have = sum(1 for v in [qty_val, price_val, amount_val] if v is not None)
        
        if have == 3:
            # 三个值都有，做交叉校验
            expected = qty_val * price_val
            if qty_val > 0 and price_val > 0 and abs(expected - amount_val) > max(0.5, amount_val * 0.01):
                # 不匹配 → 标记为可疑，不修改
                name = str(row.get(0, ''))[:15]
                log(f"  ⚠️ 可疑: {name} | 数量={qty_val}×单价={price_val}={expected:.2f}, 但金额={amount_val}")
                suspicious.add(i)
        
        elif have == 2:
            # 只缺一个值，安全地补充
            if qty_val is None and price_val and amount_val and price_val > 0:
                qty = amount_val / price_val
                row[2] = str(int(round(qty))) if abs(qty - round(qty)) < 0.02 else f"{qty:.2f}"
                log(f"  🔧 补充数量: → {row[2]}")
                fixed_count += 1
            elif price_val is None and qty_val and amount_val and qty_val > 0:
                price = amount_val / qty_val
                row[4] = str(int(price)) if price == int(price) else f"{price:.2f}"
                log(f"  🔧 补充单价: → {row[4]}")
                fixed_count += 1
            elif amount_val is None and qty_val and price_val:
                amount = qty_val * price_val
                row[5] = str(int(amount)) if amount == int(amount) else f"{amount:.2f}"
                log(f"  🔧 补充金额: → {row[5]}")
                fixed_count += 1
        
        elif have <= 1:
            # 缺太多数值，标记可疑
            if any(v is not None and v > 0 for v in [qty_val, price_val, amount_val]):
                suspicious.add(i)
    
    return suspicious, fixed_count


def create_excel(rows, header, output_path, contract_info=None, suspicious_rows=None):
    """生成带格式的 Excel，可疑行用黄色高亮标记"""
    wb = Workbook()
    ws = wb.active
    ws.title = "购销合同"

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    data_font = Font(name='微软雅黑', size=10)
    alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    warn_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色高亮
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(vertical='center', wrap_text=True)

    if suspicious_rows is None:
        suspicious_rows = set()

    cur = 1
    num_cols = len(header)

    # 合同信息
    if contract_info:
        for txt, font in [
            (contract_info.get('company', ''), Font(name='微软雅黑', size=16, bold=True)),
            ('购 销 合 同', Font(name='微软雅黑', size=14, bold=True)),
        ]:
            ws.merge_cells(f'A{cur}:{get_column_letter(num_cols)}{cur}')
            c = ws.cell(row=cur, column=1, value=txt)
            c.font = font
            c.alignment = center
            cur += 1

        for key, label in [('contract_no', '合同编号'), ('date', '日期'),
                           ('supplier', '供方'), ('buyer', '需方')]:
            val = contract_info.get(key, '')
            if val:
                ws.merge_cells(f'A{cur}:{get_column_letter(num_cols)}{cur}')
                ws.cell(row=cur, column=1,
                        value=f'{label}：{val}').font = Font(name='微软雅黑', size=11)
                cur += 1
        cur += 1

    # 表头
    for ci, name in enumerate(header, 1):
        c = ws.cell(row=cur, column=ci, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    
    # 合并 A 和 B 列表头 (对应物品名称)
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=2)
    header_row = cur
    cur += 1

    # 数据
    for ri, row in enumerate(rows):
        is_suspicious = ri in suspicious_rows
        for ci in range(num_cols):
            val = row.get(ci, '') if isinstance(row, dict) else (row[ci] if ci < len(row) else '')
            c = ws.cell(row=cur, column=ci + 1, value=val)
            c.font = data_font
            c.border = border

            if ci in (2, 3, 4, 5) and val:
                try:
                    num = float(str(val).replace(' ', '').replace('，', '').replace(',', ''))
                    if num.is_integer():
                        c.value = int(num)
                        c.number_format = '0'
                    else:
                        c.value = num
                        c.number_format = '0.##'
                    c.alignment = right if ci >= 4 else center
                except ValueError:
                    c.alignment = center if ci in (2, 3) else left
            else:
                c.alignment = left

            # 可疑行：黄色高亮（优先级高于斑马纹）
            if is_suspicious:
                c.fill = warn_fill
            elif ri % 2 == 1:
                c.fill = alt_fill
        cur += 1

    for i, w in enumerate([14, 40, 8, 6, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f'A{header_row + 1}'
    wb.save(output_path)


def process_pdf(pdf_path, output_path, dpi=300, engine='paddle', log_callback=None, progress_callback=None):
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    log(f"\n{'=' * 55}")
    log(f"  PDF 表格转 Excel 工具")
    log(f"{'=' * 55}")
    log(f"\n📄 输入：{pdf_path}")
    log(f"📊 输出：{output_path}")

    if not os.path.exists(pdf_path):
        log(f"\n❌ 找不到文件 '{pdf_path}'")
        raise FileNotFoundError(f"找不到文件 '{pdf_path}'")

    if engine == 'rapid':
        log("\n⚡ 初始化极速模型 (RapidOCR)...")
        ocr = RapidOCR(text_score=0.1, box_score_thresh=0.1)
    else:
        log("\n🎯 初始化高精度模型 (PaddleOCR)...")
        ocr = PaddleOCR(
            use_angle_cls=False,
            lang='ch',
            show_log=False
        )
    scale = dpi / 200.0  # 相对 200 DPI 的缩放比

    # 默认列边界（基于 200 DPI）
    # 列: 物品名称 | 规格型号 | 数量 | 单位 | 含税单价 | 金额 | 备注
    col_bounds_200 = [0, 250, 750, 900, 1000, 1200, 1360, 1600]

    # PDF 转图片 + OCR
    log(f"\n📷 步骤 1/3：PDF 转图片 (DPI={dpi})...")
    if progress_callback: progress_callback(0.1)
    page_data = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            tmp = f"_tmp_page_{i}.png"
            page.to_image(resolution=dpi).save(tmp)
            log(f"  ✓ 第 {i + 1}/{total_pages} 页")
            page_data.append(tmp)
            if progress_callback: progress_callback(0.1 + 0.3 * ((i + 1) / total_pages))

    # 动态检测列边界
    log(f"\n🔍 分析表头以动态计算列边界...")
    if page_data:
        first_page_img = page_data[0]
        result = full_page_ocr(ocr, first_page_img, engine_type=engine)
        if result:
            headers = {}
            for item in result:
                x_center = item['x'] / scale
                y_center = item['y'] / scale
                text = item['text']
                if y_center > 1000: continue
                for k in ['物品名称', '规格', '数量', '单位', '单价', '金额', '备注']:
                    if k in text and k not in headers:
                        headers[k] = x_center
            
            log(f"  检测到表头: {list(headers.keys())}")
            if '数量' in headers and '单价' in headers and '金额' in headers:
                x_qty = headers['数量']
                x_unit = headers.get('单位', x_qty + 100)
                x_price = headers['单价']
                x_amount = headers['金额']
                x_remark = headers.get('备注', x_amount + 200)

                b0 = 0
                b1 = headers.get('规格', headers.get('物品名称', 0) + 200) if '规格' in headers else 250
                b2 = x_qty - 60
                b3 = (x_qty + x_unit) / 2
                b4 = (x_unit + x_price) / 2
                b5 = (x_price + x_amount) / 2
                b6 = (x_amount + x_remark) / 2
                b7 = 10000
                col_bounds_200 = sorted([b0, b1, b2, b3, b4, b5, b6, b7])
                log(f"  成功应用动态边界！")
            else:
                log(f"  未找齐关键表头，使用默认边界。")

    log(f"\n🔍 步骤 2/3：OCR + 表格提取...")
    all_rows = []
    contract_info = {}
    header = ['物品名称', '', '数量', '单位', '含税单价', '金额', '备注']
    col_bounds = [int(x * scale) for x in col_bounds_200]

    for pi, img_path in enumerate(page_data):
        log(f"\n  --- 第 {pi + 1} 页 ---")

        # 检测水平线
        h_lines = detect_h_lines(img_path, min_len=int(60 * scale))
        log(f"  水平线：{len(h_lines)} 条")

        # 整页 OCR
        items = full_page_ocr(ocr, img_path, engine_type=engine)
        log(f"  OCR 文本块：{len(items)} 个")

        # 提取合同信息（第一页）
        if pi == 0 and h_lines:
            table_start_y = 0
            for y in h_lines:
                if y > int(280 * scale):
                    table_start_y = y
                    break
            for item in items:
                t = item['text']
                if item['y'] < table_start_y:
                    if re.search(r'A\d{8,}', t):
                        m = re.search(r'A\d+', t)
                        contract_info['contract_no'] = m.group()
                    if '日期' in t and '/' in t:
                        contract_info['date'] = t.replace('日期：', '').replace('日期:', '').strip()
                    if '供方' in t:
                        contract_info['supplier'] = t.replace('供方：', '').replace('供方:', '').strip()
                    if '需方' in t:
                        contract_info['buyer'] = t.replace('需方：', '').replace('需方:', '').strip()
                    if '有限公司' in t and '供方' not in t and '需方' not in t:
                        if len(t) > len(contract_info.get('company', '')):
                            contract_info['company'] = t

        # 分配到行
        row_groups = assign_to_rows(items, h_lines)

        # 确定表格数据行范围
        if pi == 0:
            # 第一页：跳过表头区域 (前几行是标题/信息)
            # 找到包含 "物品名称" 的行，下一行开始是数据
            header_row_idx = None
            for ridx, ritems in row_groups.items():
                texts = ' '.join(it['text'] for it in ritems)
                if '物品名称' in texts or '含税单价' in texts:
                    header_row_idx = ridx
                    break
            data_start = (header_row_idx + 1) if header_row_idx is not None else 2
        else:
            data_start = 0

        # 转换为表格行
        page_rows = []
        for ridx in sorted(row_groups.keys()):
            if ridx < data_start:
                continue
            row_items = row_groups[ridx]

            # 跳过页脚和合同条款
            texts = ' '.join(it['text'] for it in row_items)
            if '第' in texts and '页' in texts:
                continue
            if any(k in texts for k in ['合计', '营计', '总计', '合汁']) and len(texts.replace(' ', '')) <= 15:
                # 保留 OCR 原本的合计行，然后停止提取（过滤掉后面的无关页脚）
                row_dict = assign_to_columns(row_items, col_bounds)
                page_rows.append(row_dict)
                break  # 合计行及其后面不再有表格数据
            # 跳过合同条款
            skip_keywords = ['总计金额', '交货地点', '运输方式', '验收标准',
                             '保修期', '结算方式', '违约责任', '解决合同',
                             '本合同', '以上产品', '单位名称', '单位地址',
                             '代理人', '电话', '传真', '开户行', '账号',
                             '需方需求', '供方负担', '签字盖章']
            if any(kw in texts for kw in skip_keywords):
                continue

            # 分配到列
            row_dict = assign_to_columns(row_items, col_bounds)

            # 只保留有意义的行（至少有2个非空单元格，且包含名称或金额数据）
            non_empty = sum(1 for v in row_dict.values() if v.strip())
            has_name = bool(row_dict.get(0, '').strip() or row_dict.get(1, '').strip())
            has_money = bool(row_dict.get(4, '').strip() or row_dict.get(5, '').strip())
            if non_empty >= 2 and (has_name or has_money):
                page_rows.append(row_dict)

        all_rows.extend(page_rows)
        log(f"  ✓ 提取 {len(page_rows)} 行数据")
        if progress_callback: progress_callback(0.4 + 0.4 * ((pi + 1) / total_pages))

    log(f"\n  📋 共 {len(all_rows)} 行")

    # 后处理：OCR 数值清洗 + 交叉校验 (数量×单价=金额)
    log(f"\n🔧 数值校验与修正...")
    suspicious_rows, fixed_count = post_process_rows(all_rows, log_func=log)
    log(f"  ✅ 补充了 {fixed_count} 处缺失数据")
    if suspicious_rows:
        log(f"  ⚠️ 发现 {len(suspicious_rows)} 行可疑数据（将在 Excel 中用黄色标记）")

    # 预览
    log(f"\n  前5行预览：")
    for i, row in enumerate(all_rows[:5]):
        cells = []
        for c in range(7):
            val = row.get(c, '')[:18]
            cells.append(f'{val:18s}')
        log(f"  {i + 1}: {'|'.join(cells)}")

    # 生成 Excel
    log(f"\n📝 步骤 3/3：生成 Excel...")
    if progress_callback: progress_callback(0.9)
    create_excel(all_rows, header, output_path, contract_info, suspicious_rows=suspicious_rows)

    # 清理
    for f in page_data:
        try:
            os.remove(f)
        except OSError:
            pass

    if progress_callback: progress_callback(1.0)
    log(f"\n{'=' * 55}")
    log(f"  ✅ 完成！输出：{os.path.abspath(output_path)}")
    log(f"{'=' * 55}\n")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Window Config ---
        self.title("PDF to Excel Converter")
        self.geometry("850x650")
        self.minsize(800, 600)
        
        # Color Theme & Appearance
        ctk.set_appearance_mode("Dark")  # Force dark mode for a sleek look
        ctk.set_default_color_theme("blue")
        
        # Define modern custom colors (Slate Dark Theme)
        self.bg_color = "#0F172A"      # slate-900
        self.card_color = "#1E293B"    # slate-800
        self.accent_color = "#3B82F6"  # blue-500
        self.text_color = "#F8FAFC"    # slate-50

        self.configure(fg_color=self.bg_color)
        
        # Grid config
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # --- Header ---
        self.header_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.header_frame.grid(row=0, column=0, padx=30, pady=(30, 10), sticky="ew")
        
        self.title_label = ctk.CTkLabel(
            self.header_frame, 
            text="📄 PDF 表格转 Excel 工具", 
            font=ctk.CTkFont(family="Microsoft YaHei", size=28, weight="bold"),
            text_color=self.accent_color
        )
        self.title_label.pack(side="left")
        
        self.subtitle_label = ctk.CTkLabel(
            self.header_frame,
            text="智能提取 · 扫描件支持 · 极速生成",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            text_color="gray60"
        )
        self.subtitle_label.pack(side="right", anchor="s", pady=10)

        # --- Main Content Area (Cards) ---
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.grid(row=1, column=0, padx=30, pady=10, sticky="ew")
        self.content_frame.grid_columnconfigure(0, weight=3)
        self.content_frame.grid_columnconfigure(1, weight=1)

        # 1. File Configuration Card
        self.file_card = ctk.CTkFrame(self.content_frame, fg_color=self.card_color, corner_radius=15)
        self.file_card.grid(row=0, column=0, padx=(0, 10), sticky="nsew")
        self.file_card.grid_columnconfigure(1, weight=1)

        self.file_card_title = ctk.CTkLabel(self.file_card, text="📁 文件设置", font=ctk.CTkFont(family="Microsoft YaHei", weight="bold", size=16), text_color=self.text_color)
        self.file_card_title.grid(row=0, column=0, columnspan=3, padx=20, pady=(15, 10), sticky="w")

        # PDF Picker
        self.pdf_btn = ctk.CTkButton(
            self.file_card, text="选择 PDF", width=100, height=36,
            font=ctk.CTkFont(family="Microsoft YaHei", weight="bold"), 
            command=self.select_pdf,
            fg_color=self.accent_color, hover_color="#2563EB"
        )
        self.pdf_btn.grid(row=1, column=0, padx=(20, 10), pady=(5, 15), sticky="w")
        
        self.pdf_path_var = ctk.StringVar()
        self.pdf_entry = ctk.CTkEntry(
            self.file_card, textvariable=self.pdf_path_var, state="readonly",
            fg_color="#334155", border_width=0, corner_radius=6, height=36,
            text_color="#CBD5E1"
        )
        self.pdf_entry.grid(row=1, column=1, padx=(0, 20), pady=(5, 15), sticky="ew")

        # Output Picker
        self.out_btn = ctk.CTkButton(
            self.file_card, text="保存路径", width=100, height=36,
            font=ctk.CTkFont(family="Microsoft YaHei", weight="bold"),
            fg_color="transparent", border_width=1, border_color="#475569",
            hover_color="#334155", text_color=self.text_color,
            command=self.select_output
        )
        self.out_btn.grid(row=2, column=0, padx=(20, 10), pady=(5, 20), sticky="w")
        
        self.out_path_var = ctk.StringVar()
        self.out_entry = ctk.CTkEntry(
            self.file_card, textvariable=self.out_path_var, state="readonly",
            fg_color="#334155", border_width=0, corner_radius=6, height=36,
            text_color="#CBD5E1"
        )
        self.out_entry.grid(row=2, column=1, padx=(0, 20), pady=(5, 20), sticky="ew")

        # 2. Settings Card
        self.settings_card = ctk.CTkFrame(self.content_frame, fg_color=self.card_color, corner_radius=15)
        self.settings_card.grid(row=0, column=1, sticky="nsew")

        self.settings_title = ctk.CTkLabel(self.settings_card, text="⚙️ 识别配置", font=ctk.CTkFont(family="Microsoft YaHei", weight="bold", size=16), text_color=self.text_color)
        self.settings_title.pack(anchor="w", padx=20, pady=(15, 10))

        self.dpi_frame = ctk.CTkFrame(self.settings_card, fg_color="transparent")
        self.dpi_frame.pack(fill="x", padx=20, pady=(5, 15))
        
        self.dpi_label = ctk.CTkLabel(self.dpi_frame, text="分辨率(DPI):", text_color="gray80")
        self.dpi_label.pack(side="left")
        
        self.dpi_entry = ctk.CTkEntry(
            self.dpi_frame, width=70, justify="center",
            fg_color="#334155", border_width=0, corner_radius=6, height=32,
            text_color="#CBD5E1"
        )
        self.dpi_entry.insert(0, "300")
        self.dpi_entry.pack(side="right")

        self.engine_frame = ctk.CTkFrame(self.settings_card, fg_color="transparent")
        self.engine_frame.pack(fill="x", padx=20, pady=(5, 15))
        
        self.engine_label = ctk.CTkLabel(self.engine_frame, text="引擎:", text_color="gray80")
        self.engine_label.pack(side="left")
        
        self.engine_var = ctk.StringVar(value="高精度模式 (Paddle)")
        self.engine_menu = ctk.CTkOptionMenu(
            self.engine_frame, variable=self.engine_var,
            values=["高精度模式 (Paddle)", "极速模式 (Rapid)"],
            width=160, fg_color="#334155", button_color="#475569",
            button_hover_color="#64748B", dropdown_fg_color="#1E293B",
            text_color="#F8FAFC", command=self.on_engine_change,
            dynamic_resizing=False
        )
        self.engine_menu.pack(side="right")
        
        # Engine Description Container (Strict size to prevent layout shift)
        self.desc_container = ctk.CTkFrame(self.settings_card, fg_color="transparent", width=280, height=25)
        self.desc_container.pack_propagate(False)
        self.desc_container.pack(anchor="w", padx=20, pady=(0, 15))

        self.engine_desc = ctk.CTkLabel(
            self.desc_container, 
            text="🎯 耗时较长（约数分钟），但识别极其精准。", 
            text_color="#3B82F6", font=ctk.CTkFont(family="Microsoft YaHei", size=12),
            anchor="w"
        )
        self.engine_desc.pack(side="left")

        # Action Button
        self.start_btn = ctk.CTkButton(
            self.settings_card, text="开始转换 🚀", 
            font=ctk.CTkFont(family="Microsoft YaHei", weight="bold", size=15),
            height=45, corner_radius=8,
            fg_color="#10B981", hover_color="#059669", # Emerald green
            text_color="#FFFFFF",
            command=self.start_conversion
        )
        self.start_btn.pack(fill="x", padx=20, pady=(10, 20), side="bottom")

        # --- Logs & Progress ---
        self.log_card = ctk.CTkFrame(self, fg_color=self.card_color, corner_radius=15)
        self.log_card.grid(row=2, column=0, padx=30, pady=(10, 30), sticky="nsew")
        self.log_card.grid_rowconfigure(0, weight=1)
        self.log_card.grid_columnconfigure(0, weight=1)

        self.log_textbox = ctk.CTkTextbox(
            self.log_card, wrap="word", 
            font=ctk.CTkFont(family="Consolas", size=13),
            fg_color="#020617", text_color="#94A3B8", corner_radius=10, border_width=0
        )
        self.log_textbox.grid(row=0, column=0, padx=15, pady=(15, 10), sticky="nsew")
        
        # Welcome message
        self.log("✨ 欢迎使用 PDF 表格转 Excel 工具！")
        self.log("✨ 请在上方选择文件后点击 [开始转换]。\n" + "-"*50)

        self.progress_bar = ctk.CTkProgressBar(
            self.log_card, height=8, corner_radius=4,
            progress_color=self.accent_color, fg_color="#334155"
        )
        self.progress_bar.grid(row=1, column=0, padx=15, pady=(0, 15), sticky="ew")
        self.progress_bar.set(0)

    def on_engine_change(self, choice):
        if "Paddle" in choice:
            self.engine_desc.configure(text="🎯 耗时较长（约数分钟），但识别极其精准。", text_color="#3B82F6")
            self.start_btn.configure(fg_color="#10B981", hover_color="#059669")
        else:
            self.engine_desc.configure(text="⚡ 速度极快（约十几秒），适合清晰或页数极多的文档。", text_color="#F59E0B")
            self.start_btn.configure(fg_color="#F59E0B", hover_color="#D97706")

    def select_pdf(self):
        file_path = filedialog.askopenfilename(
            title="选择 PDF 文件",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            self.pdf_path_var.set(file_path)
            default_out = str(Path(file_path).with_suffix('.xlsx'))
            self.out_path_var.set(default_out)

    def select_output(self):
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.out_path_var.set(file_path)

    def log(self, message):
        self.log_textbox.insert(tk.END, message + "\n")
        self.log_textbox.see(tk.END)

    def update_progress(self, value):
        self.progress_bar.set(value)

    def start_conversion(self):
        pdf_path = self.pdf_path_var.get()
        out_path = self.out_path_var.get()
        try:
            dpi = int(self.dpi_entry.get())
        except ValueError:
            self.log("❌ DPI 必须是整数！")
            return

        if not pdf_path:
            self.log("❌ 请先选择 PDF 文件！")
            return
        if not out_path:
            self.log("❌ 请选择输出路径！")
            return

        self.start_btn.configure(state="disabled", text="转换中...", fg_color="gray40")
        self.progress_bar.set(0)
        self.log_textbox.delete("1.0", tk.END)
        self.log(f"🚀 开始任务，文件: {Path(pdf_path).name}\n" + "-"*50)

        # 在子线程中执行
        engine_choice = self.engine_var.get()
        engine = 'rapid' if 'Rapid' in engine_choice else 'paddle'
        thread = threading.Thread(target=self.run_conversion, args=(pdf_path, out_path, dpi, engine), daemon=True)
        thread.start()

    def run_conversion(self, pdf_path, out_path, dpi, engine):
        try:
            def thread_safe_log(msg):
                self.after(0, self.log, msg)
            
            def thread_safe_progress(val):
                self.after(0, self.update_progress, val)

            process_pdf(pdf_path, out_path, dpi=dpi, engine=engine, log_callback=thread_safe_log, progress_callback=thread_safe_progress)
        except Exception as e:
            self.after(0, self.log, f"\n❌ 发生错误: {str(e)}")
        finally:
            self.after(0, lambda: self.start_btn.configure(state="normal", text="开始转换 🚀", fg_color="#10B981"))

def main():
    if len(sys.argv) > 1:
        # 命令行模式
        parser = argparse.ArgumentParser(description='PDF 表格转 Excel（支持扫描件）')
        parser.add_argument('pdf_file', nargs='?', default='Scan.pdf', help='PDF 文件路径')
        parser.add_argument('-o', '--output', default=None, help='输出 Excel 文件路径')
        parser.add_argument('--dpi', type=int, default=300, help='OCR 分辨率（默认300）')
        parser.add_argument('--engine', type=str, choices=['rapid', 'paddle'], default='paddle', help='OCR 引擎 (默认 paddle)')
        args = parser.parse_args()
        output = args.output or f"{Path(args.pdf_file).stem}.xlsx"
        process_pdf(args.pdf_file, output, args.dpi, engine=args.engine)
    else:
        # GUI 模式
        app = App()
        app.mainloop()

if __name__ == '__main__':
    main()
